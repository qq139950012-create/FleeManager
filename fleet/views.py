import qrcode
import json
import openpyxl
from io import BytesIO
from datetime import timedelta, date, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils import timezone
from django.http import HttpResponse, HttpResponseForbidden
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.db import transaction
from django.db.models import Sum, Count, Q, Max, Min, DecimalField, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponseForbidden

# 本地模块引用
from .models import *
from .forms import *

# ==========================================
# 0. 全局配置
# ==========================================
FAULT_MAPPING = {
    'tire': '轮胎故障',
    'engine': '发动机/异响',
    'battery': '电路/电瓶',
    'body': '车身外观',
    'light': '灯光系统',
    'other': '其他问题'
}

# ==========================================
# 1. 基础路由与主页
# ==========================================
@login_required(login_url='/accounts/login/')
def home(request):
    try:
        employee = request.user.employee
    except Exception:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        return render(request, 'fleet/home.html', {'error_msg': '无员工档案'})

    if employee.role == 'admin':
        return redirect('admin_dashboard')
    if employee.role == 'leader':
        return redirect('leader_dashboard')
    if employee.role == 'dispatcher':
        return redirect('dispatch_dashboard')
    if employee.role == 'repairman':
        return redirect('repair_dashboard')

    my_tasks = WorkTask.objects.filter(
        assigned_drivers=employee, status='doing'
    ).order_by('-create_at')


    current_vehicle = Vehicle.objects.filter(
        status='working',
        current_driver=employee
    ).first()

    idle_vehicles = Vehicle.objects.filter(status='idle').order_by('vehicle_id')

    return render(request, 'fleet/home.html', {
        'driver': employee,
        'current_vehicle': current_vehicle,
        'idle_vehicles': idle_vehicles,
        'my_tasks': my_tasks
    })


driver_dashboard = home

# ==========================================
# 2. 核心业务：扫码、绑定、解绑
# ==========================================
@login_required
def driver_scan_vehicle(request):
    return redirect('home')


@login_required
def vehicle_start(request, vehicle_id):
    e = request.user.employee

    # ✅ 必须先上班才能绑定车（后端强制）
    if e.work_status != 'on_duty':
        messages.error(request, "请先点击【上班】后再扫码绑定车辆")
        return redirect('home')

    v = get_object_or_404(Vehicle, id=vehicle_id)

    # ✅ 已经有工作中的车，就不允许再绑第二台
    current_vehicle = Vehicle.objects.filter(status='working', current_driver=e).first()
    if current_vehicle and current_vehicle.id != v.id:
        messages.error(request, f"你当前已绑定车辆：{current_vehicle.plate_number}，请先解绑再绑定其他车辆")
        return redirect('home')

    if request.method == 'POST':
        # ✅ 关键：带 instance，避免 fields='__all__' 时缺字段报错
        form = StartWorkForm(request.POST, instance=v)
        if form.is_valid():
            mileage = form.cleaned_data.get('current_mileage')

            # ✅ 出车里程不能小于系统里程（后端强制）
            if mileage is not None and mileage < (v.current_mileage or 0):
                messages.error(request, f"出车里程不能小于当前系统里程（当前：{v.current_mileage or 0} km）")
                return redirect('vehicle_start', vehicle_id=v.id)

            # ✅ 更新车辆里程（只有填了才更新）
            if mileage is not None:
                v.current_mileage = mileage

            v.status = 'working'
            v.current_driver = e
            v.save()

            messages.success(request, f"成功绑定车辆：{v.plate_number}")
            return redirect('home')

        messages.error(request, "表单验证失败，请检查里程输入")
    else:
        form = StartWorkForm(instance=v)

    return render(request, 'fleet/action_start.html', {'vehicle': v, 'form': form})


@login_required
def vehicle_end(request, vehicle_id):
    me = request.user.employee
    v = get_object_or_404(Vehicle, id=vehicle_id)

    # 车辆本来就没绑定 / 不在工作中
    if v.status != 'working' or not v.current_driver:
        messages.info(request, "车辆当前不是工作状态，无需解绑")
        return redirect('home')

    driver = v.current_driver  # 当前绑定司机

    # ✅ 权限判断：本人 / 同班班长 / 管理员
    is_self = (driver.id == me.id)
    is_leader = (me.role in ['leader', 'team_leader']) and (me.team and driver.team == me.team)
    is_admin = (me.role == 'admin') or request.user.is_superuser

    if not (is_self or is_leader or is_admin):
        messages.error(request, "你无权解绑这台车（仅司机本人/同班班长/管理员可操作）")
        return redirect('home')

    # ✅ POST：执行收车 + 更新里程
    if request.method == 'POST':
        form = EndWorkForm(request.POST, instance=v)
        if not form.is_valid():
            messages.error(request, "表单验证失败，请检查里程输入")
            return redirect('vehicle_end', vehicle_id=v.id)

        mileage = form.cleaned_data.get('current_mileage')

        # ✅ 里程不能倒退
        if mileage is not None and mileage < (v.current_mileage or 0):
            messages.error(request, f"收车里程不能小于当前里程（当前：{v.current_mileage or 0} km）")
            return redirect('vehicle_end', vehicle_id=v.id)

        # ✅ 更新里程
        if mileage is not None:
            v.current_mileage = mileage

        # ✅ 解绑收车
        v.status = 'idle'
        v.current_driver = None
        v.save()

        # ✅ 返回提示
        if is_self:
            messages.success(request, "已收车，车辆恢复空闲")
        elif is_leader:
            messages.success(request, f"已代 {driver.name} 收车，车辆恢复空闲")
        else:
            messages.success(request, f"管理员已代 {driver.name} 收车，车辆恢复空闲")

        return redirect('home')

    # ✅ GET：打开收车页面（默认显示当前里程）
    form = EndWorkForm(initial={'current_mileage': v.current_mileage}, instance=v)
    return render(request, 'fleet/action_end.html', {'vehicle': v, 'form': form})


# ==========================================
# 3. 司机日常运营
# ==========================================
@login_required
def add_record(request, vehicle_id, type):
    v = get_object_or_404(Vehicle, id=vehicle_id)
    if request.method == 'POST':
        form = OperationForm(request.POST, request.FILES)
        if form.is_valid():
            r = form.save(commit=False)
            r.vehicle = v
            r.driver = request.user.employee
            r.record_type = type
            if type == 'fuel' and form.cleaned_data.get('location_type'):
                r.remark = f"[{form.cleaned_data.get('location_type')}] {r.remark or ''}"
            r.save()
            if r.mileage and r.mileage > v.current_mileage:
                v.current_mileage = r.mileage
                v.save()
            messages.success(request, "记录已提交")
            return redirect('home')
    else:
        form = OperationForm(initial={'record_type': type})
    return render(request, 'fleet/driver_form.html', {'form': form, 'vehicle': v})


@login_required
def add_repair(request, vehicle_id):
    v = get_object_or_404(Vehicle, id=vehicle_id)
    if request.method == 'POST':
        form = RepairForm(request.POST, request.FILES)
        if form.is_valid():
            r = form.save(commit=False)
            r.vehicle = v
            r.driver = request.user.employee
            r.fault_type = form.cleaned_data.get('fault_category', 'other')
            r.status = 'pending'
            r.save()
            v.status = 'repairing'
            v.save()
            messages.warning(request, "故障已上报，车辆已锁定")
            return redirect('home')
    else:
        form = RepairForm()
    return render(request, 'fleet/driver_repair.html', {'form': form, 'vehicle': v})


# ==========================================
# 4. 二维码与工具
# ==========================================
@login_required
def generate_qrcode(request, vehicle_id):
    relative_path = reverse('vehicle_start', args=[vehicle_id])
    real_domain = "https://7777777l.com"
    scan_url = f"{real_domain}{relative_path}"
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=15,
        border=4
    )
    qr.add_data(scan_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    return HttpResponse(buf.getvalue(), content_type="image/png")


@login_required
def toggle_work_status(request):
    e = request.user.employee

    # ✅ 如果准备下班：必须先解绑车
    if e.work_status == 'on_duty':
        current_vehicle = Vehicle.objects.filter(status='working', current_driver=e).first()
        if current_vehicle:
            messages.error(request, f"下班前请先解绑车辆：{current_vehicle.plate_number}")
            return redirect('home')

        e.work_status = 'off_duty'
        e.save()
        messages.success(request, "已下班")
        return redirect('home')

    # ✅ 上班
    e.work_status = 'on_duty'
    e.save()
    messages.success(request, "已上班")
    return redirect('home')


@login_required
def sign_out(request):
    logout(request)
    return redirect('login')


# ==========================================
# 5. 管理员后台
# ==========================================
@login_required
def admin_dashboard(request):
    try:
        if request.user.employee.role != 'admin' and not request.user.is_superuser:
            return redirect('home')
    except Exception:
        return redirect('home')

    context = {
        'employees': Employee.objects.all().order_by('role'),
        'vehicles': Vehicle.objects.all().order_by('vehicle_id'),
        'inventory': InventoryItem.objects.all(),
        'me': request.user.employee,
    }
    return render(request, 'fleet/admin_dashboard.html', context)


@login_required
def admin_vehicle_manage(request, v_id=None):
    if request.user.employee.role != 'admin':
        return redirect('home')
    v = get_object_or_404(Vehicle, id=v_id) if v_id else None
    if request.method == 'POST':
        if 'delete' in request.POST and v:
            v.delete()
            return redirect('admin_dashboard')
        form = AdminVehicleForm(request.POST, instance=v)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    else:
        form = AdminVehicleForm(instance=v)
    return render(request, 'fleet/admin_form.html', {'form': form, 'title': '车辆管理', 'vehicle': v})


@login_required
def admin_employee_manage(request, emp_id=None):
    if request.user.employee.role != 'admin':
        return redirect('home')
    target = get_object_or_404(Employee, id=emp_id) if emp_id else None
    if request.method == 'POST':
        form = AdminEmployeeForm(request.POST, instance=target)
        if form.is_valid():
            emp = form.save(commit=False)
            username = form.cleaned_data.get('username')
            if not target:
                user = User.objects.create_user(username=username, password="123")
                emp.user = user
            emp.save()
            return redirect('admin_dashboard')
    else:
        form = AdminEmployeeForm(instance=target)
    return render(request, 'fleet/admin_form.html', {'form': form, 'title': '员工管理'})


@login_required
def admin_bonus_list(request):
    return render(request, 'fleet/admin_bonus_list.html', {
        'batches': BonusBatch.objects.all().order_by('-created_at')
    })


@login_required
def admin_bonus_create(request):
    if request.method == 'POST':
        f = AdminBonusCreateForm(request.POST)
        if f.is_valid():
            b = f.save(commit=False)
            leader = Employee.objects.filter(role='team_leader').first()
            if leader:
                b.leader = leader
                b.status = 'pending'
                b.save()
                messages.success(request, "奖金包已创建")
            else:
                messages.error(request, "未找到班长")
            return redirect('admin_bonus_list')
    return render(request, 'fleet/admin_form.html', {'form': AdminBonusCreateForm(), 'title': '发放奖金包'})


def export_excel(headers, rows, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


@login_required
def admin_vehicle_export(request):
    rows = [[v.vehicle_id, v.plate_number, v.brand_model, v.get_status_display()] for v in Vehicle.objects.all()]
    return export_excel(['自编号', '车牌', '型号', '状态'], rows, f"车辆_{date.today()}")


@login_required
def admin_employee_export(request):
    rows = [[e.name, e.phone, e.get_role_display(), e.team] for e in Employee.objects.all()]
    return export_excel(['姓名', '手机', '角色', '班组'], rows, f"员工_{date.today()}")


@login_required
def admin_inventory_export(request):
    rows = [[i.name, i.spec, i.unit, i.stock] for i in InventoryItem.objects.all()]
    return export_excel(['名称', '规格', '单位', '库存'], rows, f"库存_{date.today()}")


@login_required
def admin_bonus_export(request):
    rows = [[b.month, b.title, b.total_amount, b.leader.name] for b in BonusBatch.objects.all()]
    return export_excel(['月份', '标题', '金额', '负责人'], rows, f"奖金_{date.today()}")


@login_required
def admin_download_template(request):
    return export_excel(['自编号', '车牌', '型号'], [['V01', '京A8888', '红旗']], "车辆模板")


@login_required
def admin_download_employee_template(request):
    return export_excel(['用户名', '姓名', '手机', '角色', '班组'], [['zs', '张三', '13800', 'driver', '一班']], "员工模板")


@login_required
def admin_download_inventory_template(request):
    return export_excel(['名称', '规格', '单位', '库存'], [['轮胎', 'R16', '个', '10']], "库存模板")


@login_required
def admin_employee_import(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(request.FILES['file'])
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        if not User.objects.filter(username=str(row[0])).exists():
                            user = User.objects.create_user(username=str(row[0]), password="123")
                            Employee.objects.create(
                                user=user, name=row[1], phone=str(row[2]), role=row[3], team=row[4]
                            )
                messages.success(request, "导入成功")
                return redirect('admin_dashboard')
            except Exception as e:
                messages.error(request, f"失败: {e}")
    return render(request, 'fleet/admin_import.html', {'form': UploadFileForm(), 'type': 'employee'})


@login_required
def admin_vehicle_import(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(request.FILES['file'])
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        Vehicle.objects.get_or_create(
                            vehicle_id=row[0],
                            defaults={'plate_number': row[1], 'brand_model': row[2]}
                        )
                messages.success(request, "导入成功")
                return redirect('admin_dashboard')
            except Exception as e:
                messages.error(request, f"失败: {e}")
    return render(request, 'fleet/admin_import.html', {'form': UploadFileForm(), 'type': 'vehicle'})


@login_required
def admin_inventory_import(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(request.FILES['file'])
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        InventoryItem.objects.create(
                            name=row[0], spec=row[1], unit=row[2], stock=row[3] or 0
                        )
                messages.success(request, "导入成功")
                return redirect('inventory_list')
            except Exception as e:
                messages.error(request, f"失败: {e}")
    return render(request, 'fleet/admin_import.html', {'form': UploadFileForm(), 'type': 'inventory'})


@login_required
def admin_inventory_manage(request, item_id=None):
    return redirect('admin_dashboard')


@login_required
def admin_bonus_report(request, batch_id):
    return redirect('admin_dashboard')


# ==========================================
# 6. 班长逻辑
# ==========================================
@login_required
def leader_dashboard(request):
    return redirect('team_dashboard')


@login_required
def team_dashboard(request):
    me = request.user.employee
    if me.role not in ['team_leader', 'leader', 'dispatcher']:
        return redirect('home')

    members = Employee.objects.filter(team=me.team).exclude(id=me.id) if me.team else Employee.objects.none()

    working_count = 0
    member_data = []
    for m in members:
        v = Vehicle.objects.filter(status='working', current_driver=m).first()
        status = 'working' if v else m.work_status
        if v:
            working_count += 1
        member_data.append({'info': m, 'vehicle': v, 'status': status})

    pending_bonuses = BonusBatch.objects.filter(leader=me, status='pending')
    stats = {
        'total': members.count(),
        'working': working_count,
        'on_duty': members.filter(work_status='on_duty').count(),
        'off_duty': members.filter(work_status='off_duty').count()
    }

    # ✅ 新增：班长要看的任务（按班组来显示，不会因为 assigned_drivers 覆盖而丢失）
    leader_tasks = WorkTask.objects.filter(
        target_team=me.team
    ).exclude(status='done').order_by('-create_at')

    return render(request, 'fleet/leader_dashboard.html', {
        'me': me,
        'stats': stats,
        'members': member_data,
        'pending_bonuses': pending_bonuses,
        'leader_tasks': leader_tasks,   # ✅ 传给模板
    })



@login_required
def leader_bonus_distribute(request, batch_id):
    me = request.user.employee
    batch = get_object_or_404(BonusBatch, id=batch_id)
    if batch.leader != me or batch.status == 'completed':
        return redirect('team_dashboard')

    members = Employee.objects.filter(team=me.team).exclude(id=me.id)

    if request.method == 'POST':
        with transaction.atomic():
            total = 0
            details = []
            for m in members:
                amt = float(request.POST.get(f'amount_{m.id}', 0))
                if amt > 0:
                    total += amt
                    details.append(BonusDetail(batch=batch, receiver=m, amount=amt))

            if abs(total - float(batch.total_amount)) > 0.1:
                messages.error(request, f"金额不匹配：已分{total} / 总额{batch.total_amount}")
            else:
                BonusDetail.objects.bulk_create(details)
                batch.status = 'completed'
                batch.distributed_at = timezone.now()
                batch.save()
                messages.success(request, "分配成功")
                return redirect('team_dashboard')

    return render(request, 'fleet/leader_bonus_distribute.html', {'batch': batch, 'members': members})


# ==========================================
# 7. 维修逻辑
# ==========================================
@login_required
def repair_dashboard(request):
    me = request.user.employee
    if me.role != 'repairman' and not request.user.is_superuser:
        return redirect('home')

    pending = MaintenanceRecord.objects.filter(status='pending').order_by('-create_at')
    my_repairs = MaintenanceRecord.objects.filter(status='repairing', repairman=me).order_by('-create_at')

    for r in pending:
        r.fault_text = FAULT_MAPPING.get(r.fault_type, '未知故障')
    for r in my_repairs:
        r.fault_text = FAULT_MAPPING.get(r.fault_type, '未知故障')

    return render(request, 'fleet/repair_dashboard.html', {
        'pending_repairs': pending,
        'my_repairs': my_repairs,
        'me': me
    })


@login_required
def repair_start(request, repair_id):
    r = get_object_or_404(MaintenanceRecord, id=repair_id)
    r.status = 'repairing'
    r.repairman = request.user.employee
    r.save()
    return redirect('repair_detail', repair_id=r.id)


@login_required
def repair_detail(request, repair_id):
    r = get_object_or_404(MaintenanceRecord, id=repair_id)
    r.fault_text = FAULT_MAPPING.get(r.fault_type, '未知故障')
    used = PartUsage.objects.filter(maintenance=r)
    parts = InventoryItem.objects.all()
    return render(request, 'fleet/repair_detail.html', {'repair': r, 'used_parts': used, 'all_parts': parts})


@login_required
def repair_add_part(request, repair_id):
    r = get_object_or_404(MaintenanceRecord, id=repair_id)
    if request.method == 'POST':
        pid = request.POST.get('part_id')
        qty = int(request.POST.get('quantity', 0))
        part = get_object_or_404(InventoryItem, id=pid)
        if part.stock >= qty:
            part.stock -= qty
            part.save()
            PartUsage.objects.create(maintenance=r, part=part, quantity=qty)
            messages.success(request, "领料成功")
        else:
            messages.error(request, "库存不足")
    return redirect('repair_detail', repair_id=r.id)


@login_required
def repair_complete(request, repair_id):
    r = get_object_or_404(MaintenanceRecord, id=repair_id)
    if request.method == 'POST':
        r.status = 'completed'
        r.finish_at = timezone.now()
        r.save()
        r.vehicle.status = 'idle'
        r.vehicle.save()
        messages.success(request, "维修完成")
        return redirect('repair_dashboard')
    return redirect('repair_detail', repair_id=r.id)


# ==========================================
# 8. 库存逻辑
# ==========================================
@login_required
def inventory_list(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_new':
            InventoryItem.objects.create(
                name=request.POST.get('name'),
                spec=request.POST.get('spec'),
                unit=request.POST.get('unit'),
                category=request.POST.get('category', 'spare'),
                stock=0
            )
            messages.success(request, "新增配件品类成功")

        elif action == 'stock_change':
            item = get_object_or_404(InventoryItem, id=request.POST.get('item_id'))
            qty = int(request.POST.get('quantity', 0))

            if request.POST.get('type') == 'in':
                item.stock += qty
                item.save()
                messages.success(request, "入库成功")
            else:
                if item.stock >= qty:
                    item.stock -= qty
                    item.save()
                    messages.success(request, "出库成功")
                else:
                    messages.error(request, "库存不足")

        return redirect('inventory_list')

    return render(request, 'fleet/inventory_list.html', {'items': InventoryItem.objects.all()})


@login_required
def inventory_add(request):
    return redirect('inventory_list')


@login_required
def inventory_action(request, item_id):
    return redirect('inventory_list')


# ==========================================
# 9. 调度与任务
# ==========================================
@login_required
def dispatch_dashboard(request):
    me = request.user.employee
    if me.role != 'dispatcher' and not request.user.is_superuser:
        return redirect('home')

    if request.method == 'POST':
        form = TaskCreateForm(request.POST)
        if form.is_valid():
            t = form.save(commit=False)
            t.created_by = me
            t.status = 'pending'
            t.save()

            # ✅ 创建后：自动下发给该班组的班长（assigned_drivers 先只放班长）
            leader = Employee.objects.filter(role='team_leader', team=t.target_team).first()
            if not leader:
                messages.error(request, f"该班组【{t.target_team}】未找到班长（team_leader），请先在后台补齐班长账号的班组")
                t.delete()
                return redirect('dispatch_dashboard')

            t.assigned_drivers.set([leader])
            messages.success(request, f"任务已下发给班长：{leader.name}")
            return redirect('dispatch_dashboard')
    else:
        form = TaskCreateForm()

    return render(request, 'fleet/dispatch_dashboard.html', {
        'vehicles': Vehicle.objects.all(),
        'form': form,
        'active_tasks': WorkTask.objects.exclude(status='done').order_by('-create_at'),
        'me': me
    })

from django.http import HttpResponseForbidden

@login_required
def dispatch_send_to_leader(request, task_id):
    """调度：把任务下发给某个班长（assigned_drivers 先只放班长，状态保持 pending）"""
    me = request.user.employee
    if me.role != 'dispatcher' and not request.user.is_superuser:
        return HttpResponseForbidden("403 Forbidden")

    t = get_object_or_404(WorkTask, id=task_id)

    if request.method != 'POST':
        return HttpResponseForbidden("403 Forbidden")

    leader_id = request.POST.get('leader_id')
    leader = get_object_or_404(Employee, id=leader_id)

    if leader.role not in ['team_leader', 'leader']:
        messages.error(request, "请选择正确的班长账号")
        return redirect('dispatch_dashboard')

    # 下发给班长：只把班长放进去，状态 pending
    t.assigned_drivers.set([leader])
    t.status = 'pending'
    t.save()

    messages.success(request, f"已下发任务给班长：{leader.name}")
    return redirect('dispatch_dashboard')


@login_required
def leader_assign_task(request, task_id):
    """班长：把任务分配给班组成员（包含自己），并把状态改为 doing"""
    me = request.user.employee
    if me.role not in ['team_leader', 'leader'] and not request.user.is_superuser:
        return HttpResponseForbidden("403 Forbidden")

    t = get_object_or_404(WorkTask, id=task_id)

    # ✅ 必须是“发给我的任务”才能分配（防止别的班长乱点）
    if not request.user.is_superuser:
        if me not in t.assigned_drivers.all():
            return HttpResponseForbidden("403 Forbidden")

    members = Employee.objects.filter(team=me.team) if me.team else Employee.objects.none()

    if request.method == 'POST':
        ids = request.POST.getlist('members')
        targets = Employee.objects.filter(id__in=ids)

# ✅ 强制包含班长自己（符合你的流程：包括自己）
targets = Employee.objects.filter(Q(id__in=ids) | Q(id=me.id))

if not targets.exists():
    messages.error(request, "请至少选择 1 个成员")
    return redirect('leader_assign_task', task_id=t.id)

t.assigned_drivers.set(targets)
t.status = 'doing'
t.save()


        messages.success(request, "任务已分配，成员端将只看到当前任务")
        return redirect('team_dashboard')

    return render(request, 'fleet/leader_task_assign.html', {
        'task': t,
        'members': members,
        'me': me,
    })


@login_required
def leader_finish_task(request, task_id):
    """班长：任务完成（done），调度端就能看到反馈"""
    me = request.user.employee
    if me.role not in ['team_leader', 'leader'] and not request.user.is_superuser:
        return HttpResponseForbidden("403 Forbidden")

    t = get_object_or_404(WorkTask, id=task_id)

    if request.method != 'POST':
        return HttpResponseForbidden("403 Forbidden")

    # ✅ 只能完成自己班组相关任务（超级管理员除外）
    if not request.user.is_superuser:
        if not t.assigned_drivers.filter(team=me.team).exists():
            return HttpResponseForbidden("403 Forbidden")

    t.status = 'done'
    t.save()
    messages.success(request, "任务已完成，已反馈给调度")
    return redirect('team_dashboard')



@login_required
def leader_force_end(request, vehicle_id):
    me = request.user.employee
    v = get_object_or_404(Vehicle, id=vehicle_id)

    # ✅ 必须 POST
    if request.method != 'POST':
        return HttpResponseForbidden("403 Forbidden")

    is_admin = (me.role == 'admin') or request.user.is_superuser
    is_leader = (me.role in ['leader', 'team_leader']) and (
        me.team and v.current_driver and v.current_driver.team == me.team
    )

    # ✅ 没权限：直接 403（彻底进不去）
    if not (is_admin or is_leader):
        return HttpResponseForbidden("403 Forbidden")

    if v.status != 'working' or not v.current_driver:
        messages.info(request, "车辆当前不是作业状态，无需代收")
        return redirect('team_dashboard')

    driver_name = getattr(v.current_driver, 'name', '该司机')
    plate = v.plate_number

    v.status = 'idle'
    v.current_driver = None
    v.save()

    messages.success(request, f"已代 {driver_name} 收车：{plate}")
    return redirect('team_dashboard')


@login_required
def assign_task_to_driver(request, task_id):
    t = get_object_or_404(WorkTask, id=task_id)
    if request.method == 'POST':
        dids = request.POST.getlist('drivers')
        t.assigned_drivers.set(Employee.objects.filter(id__in=dids))
        t.status = 'assigned'
        t.save()
    return redirect('dispatch_dashboard')


@login_required
def complete_task(request, task_id):
    t = get_object_or_404(WorkTask, id=task_id)
    t.status = 'completed'
    t.save()
    return redirect('home')


@login_required
def leader_complete_task(request, task_id):
    return complete_task(request, task_id)


@login_required
def leader_reset_task(request, task_id):
    t = get_object_or_404(WorkTask, id=task_id)
    t.status = 'pending'
    t.assigned_drivers.clear()
    t.save()
    return redirect('dispatch_dashboard')


@login_required
def tv_dashboard(request):
    stats = {
        'total': Vehicle.objects.count(),
        'working': Vehicle.objects.filter(status='working').count(),
        'repair': Vehicle.objects.filter(status='repairing').count(),
        'idle': Vehicle.objects.filter(status='idle').count()
    }
    return render(request, 'fleet/tv_dashboard.html', {
        'stats': stats,
        'recent_ops': OperationRecord.objects.order_by('-create_at')[:8],
        'recent_repairs': MaintenanceRecord.objects.order_by('-create_at')[:5]
    })


@login_required
def vehicle_dashboard(request, vehicle_id):
    return redirect('vehicle_start', vehicle_id=vehicle_id)
